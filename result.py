from openpyxl import Workbook,load_workbook
wb=load_workbook('test1.xlsx')
ws=wb.active
max_row = ws.max_row
n = 0

for row in range(2,max_row + 1):
    id = ws['A'+str(row)].value
    rate = ws['B'+str(row)].value
    hours = ws['C'+str(row)].value
    if(type(rate) != str and type(hours) != str):
        salary = rate*hours
        ws['D' + str(row)].value = salary
    if(salary >= 3000):
        #print("People id " + str(id) + " His salary is " + str(salary))
        n=n+1
print(n)
wb.close()  